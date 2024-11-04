from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.urls import reverse
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
from django.contrib.auth.models import User
from .models import Tour, Booking, Comment
from .forms import SignUpForm, ProfileForm, TourForm, CommentForm
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse
from .models import UserProfile
import json
from datetime import datetime
from django.db.models import Sum
# from django.core.mail import send_mail
from .email_utils import send_confirmation_email
from django.conf import settings
from django.http import HttpResponse
from openpyxl.chart import BarChart, Reference
from openpyxl import Workbook
from django.core.mail import send_mail
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import openpyxl

def home(request):
    tours_list = Tour.objects.all()  # Get all tours

    # Implement pagination
    paginator = Paginator(tours_list, 6)  # Show 5 tours per page
    page_number = request.GET.get('page')  # Get the current page number from the query parameters
    tours = paginator.get_page(page_number)  # Get the tours for the current page

    return render(request, 'tour/home.html', {
        'tours': tours,
        'query': request.GET.get('q', ''),  # Pass the query if applicable
        'sort_by': request.GET.get('sort', 'default'),  # Pass sorting if applicable
    })

def tour_detail(request, tour_id):
    tour = get_object_or_404(Tour, id=tour_id)
    comments = Comment.objects.filter(tour=tour)
    
    # Prepare the data for the map
    tour_data = {
        'latitude': tour.latitude,
        'longitude': tour.longitude,
        'title': tour.title,
        'description': tour.description
    }
    # Convert to JSON and pass it to the template
    tour_json = json.dumps(tour_data)

    context = {
        'tour': tour,
        'tour_json': tour_json,
        'comments': comments,
    }
    return render(request, 'tour/tour_detail.html', context)

def user_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)

            # Check if the user has a UserProfile
            user_profile, created = UserProfile.objects.get_or_create(user=user)
            if created:
                return redirect('create_profile_view')  # Redirect to profile creation page
            
            # Check if the user is a superuser
            if user.is_superuser:
                return redirect('admin_dashboard')  # Redirect to admin dashboard for superusers
            else:
                return redirect('home')  # Redirect regular users to home

        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'tour/login.html')

def signup(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])
            user.save()
            messages.success(request, 'Account created successfully!')
            return redirect('login')
    else:
        form = SignUpForm()
    return render(request, 'tour/signup.html', {'form': form})

@login_required
def book_tour(request, tour_id):
    # Find the tour the user wants to book
    tour = get_object_or_404(Tour, id=tour_id)
    
    if request.method == 'POST':
        # Placeholder for actual payment processing logic
        payment_successful = process_payment()  # Replace this with the actual payment function
        
        # If payment is successful, redirect to complete payment
        if payment_successful:
            # Send confirmation email to the user
            user = request.user  # Get the logged-in user
            tour_details = f'Tour Name: {tour.name}\nDetails: {tour.details}'  # Customize as needed
            send_confirmation_email(user, tour_details)  # Send the email

            # Redirect to payment completion page
            return redirect('complete_payment', tour_id=tour_id)
        else:
            messages.error(request, 'Payment failed. Please try again.')
            return redirect('tour_detail', tour_id=tour_id)
    
    # Render booking confirmation or initial booking page
    return render(request, 'tour/confirm_booking.html', {'tour': tour})

def success_page(request, tour_id):
    return render(request, 'tour/sucess.html', {'tour_id': tour_id})  # Pass tour_id to the template


def success(request, tour_id):
    # Logic for the success page using tour_id
    return render(request, 'tour/sucess.html', {'tour_id': tour_id})

@login_required
def my_bookings(request):
    # Retrieve all bookings for the current user
    bookings = Booking.objects.filter(user=request.user)
    
    # Directly render the 'my_bookings' page with all the bookings
    return render(request, 'tour/my_bookings.html', {'bookings': bookings})


def cancel_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)
    booking.delete()  # Assuming canceling means deleting the booking
    return redirect('my_bookings') 

def user_logout(request):
    logout(request)
    return redirect('home')

@user_passes_test(lambda u: u.is_superuser)
def admin_dashboard(request):
    total_tours = Tour.objects.count()
    total_bookings = Booking.objects.count()
    total_users = UserProfile.objects.count()

    # Calculate total revenue
    total_revenue = Booking.objects.aggregate(total=Sum('total_amount'))['total'] or 0

    # Get current year to ensure the month filtering works correctly for the current year
    current_year = datetime.now().year

    # Monthly revenue calculations
    revenue_by_month = [
        Booking.objects.filter(booked_on__year=current_year, booked_on__month=i).aggregate(total=Sum('total_amount'))['total'] or 0 
        for i in range(1, 13)
    ]

    # Updated to use 'booked_on' field for filtering bookings by month
    bookings_by_month = [
        Booking.objects.filter(booked_on__year=current_year, booked_on__month=i).count() 
        for i in range(1, 13)
    ]

    # Assuming Tour model has a field like 'created_on' or similar for filtering by month
    tours_by_month = [
        Tour.objects.filter(created_on__year=current_year, created_on__month=i).count() 
        for i in range(1, 13)
    ]

    # Get recent bookings (last 5 for example)
    recent_bookings = Booking.objects.order_by('-booked_on')[:5]

    context = {
        'total_tours': total_tours,
        'total_bookings': total_bookings,
        'total_users': total_users,
        'total_revenue': total_revenue,
        'revenue_by_month': revenue_by_month,
        'bookings_by_month': bookings_by_month,
        'tours_by_month': tours_by_month,
        'recent_bookings': recent_bookings,
    }

    return render(request, 'tour/dashboard.html', context)

def tour_list(request):
    # Get search query and sorting options from the request
    query = request.GET.get('q', '')  # Default to an empty string if no search query
    sort_by = request.GET.get('sort', 'created_on')  # Default sorting by 'created_at'
    
    # Filter tours based on the search query
    tours = Tour.objects.filter(
        Q(title__icontains=query) | Q(description__icontains=query)
    )
    
    # Sort tours based on the selected option
    if sort_by == 'price':
        tours = tours.order_by('price')
    else:
        tours = tours.order_by('-created_on')  # Default sorting by newest

    # Pagination setup
    paginator = Paginator(tours, 3)  # Show 6 tours per page
    page_number = request.GET.get('page')
    tours_page = paginator.get_page(page_number)
    
    # Pass relevant data to the template
    context = {
        'tours': tours_page,  # Paginated tours
        'query': query,  # Search query string
        'sort_by': sort_by,  # Selected sort option
    }
    
    return render(request, 'tour/tour_list.html', context)

def tour_create(request):
    if request.method == 'POST':
        form = TourForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tour added successfully!')
            return redirect('tour_list')
    else:
        form = TourForm()  # Empty form for creation
    return render(request, 'tour/tour_form.html', {'form': form})

def tour_edit(request, pk):
    tour = get_object_or_404(Tour, pk=pk)
    if request.method == 'POST':
        form = TourForm(request.POST, request.FILES, instance=tour)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tour updated successfully!')
            return redirect('tour_list')
    else:
        form = TourForm(instance=tour)  # Existing data for editing
    return render(request, 'tour/tour_form.html', {'form': form})


def tour_delete(request, pk):
    tour = get_object_or_404(Tour, pk=pk)
    tour.delete()
    messages.success(request, 'Tour deleted successfully!')
    return redirect('tour_list')


def booking_list(request):
    bookings = Booking.objects.all()
    return render(request, 'tour/booking_list.html', {'bookings': bookings})

@login_required
def booking_delete(request, pk):
    # Retrieve the booking or raise a 404 error if not found
    booking = get_object_or_404(Booking, pk=pk)
    
    # Delete the booking
    booking.delete()
    messages.success(request, 'Booking deleted successfully!')
    
    # Redirect back to the referring page or to the booking list if referer is not available
    referer = request.META.get('HTTP_REFERER')
    if referer:
        return redirect(referer)
    else:
        return redirect('tour/booking_list') 

def user_list(request):
    users = User.objects.all()
    return render(request, 'tour/user_list.html', {'users': users})

def user_delete(request, pk):
    user = get_object_or_404(User, pk=pk)
    user.delete()
    messages.success(request, 'User deleted successfully!')
    return redirect('user_list')


@login_required
def profile_view(request):
    if request.method == 'POST':
        user = request.user
        user.username = request.POST.get('username')
        user.email = request.POST.get('email')

        # Update user profile fields
        user_profile = user.userprofile
        user_profile.phone_number = request.POST.get('phone_number')
        user_profile.address = request.POST.get('address')

        # Handle avatar upload
        if request.FILES.get('avatar'):
            user_profile.avatar = request.FILES.get('avatar')

        # Save the user and profile
        user.save()
        user_profile.save()

        messages.success(request, 'Profile updated successfully!')
        return redirect('profile')  # Redirect to the profile page or any other page

    return render(request, 'tour/profile.html')


def edit_profile_view(request):
    # Logic for editing profile
    return render(request, 'tour/edit_profile.html') 

@login_required
def add_comment(request, tour_id):
    tour = get_object_or_404(Tour, id=tour_id)
    if request.method == 'POST':
        comment_text = request.POST.get('comment')
        rating = request.POST.get('rating')
        
        # Create a new comment instance
        Comment.objects.create(tour=tour, user=request.user, text=comment_text, rating=rating)

        # Add a success message
        messages.success(request, 'Your comment has been added successfully!')

        return redirect('tour_detail', tour_id=tour.id)

    return redirect('tour_detail', tour_id=tour.id)


def payment_page(request, tour_id):
    tour = get_object_or_404(Tour, id=tour_id)
    return render(request, 'tour/payment.html', {'tour': tour})

@login_required
def complete_payment(request, tour_id):
    # Get the tour for which payment was made
    tour = get_object_or_404(Tour, id=tour_id)

    # Confirm that a booking does not already exist
    booking, created = Booking.objects.get_or_create(
        user=request.user,
        tour=tour,
        defaults={
            'total_amount': tour.price,  # Assuming `tour.price` exists
            'payment_status': True
        }
    )
    
    if created:
        # Booking is successfully created
        messages.success(request, 'Booking confirmed successfully!')
    else:
        # If booking already exists
        messages.info(request, 'You have already booked this tour.')

    # Redirect to the "My Bookings" page to view the booking
    return redirect('my_bookings')



def tour_map(request):
    # Retrieve all tours with required fields
    tours = Tour.objects.values('title', 'description', 'latitude', 'longitude')
    tours_json = json.dumps(list(tours))  # Convert to JSON format
    return render(request, 'tours/tour_list.html', {'tours_json': tours_json})

def send_confirmation_email(user, tour_details):
    # Get the user's email directly from the User model
    user_email = user.email  # This is the email field from the User model

    subject = 'Booking Confirmation'
    message = f'Thank you for booking your tour! Here are the details:\n{tour_details}'
    email_from = settings.EMAIL_HOST_USER
    recipient_list = [user_email]

    try:
        send_mail(subject, message, email_from, recipient_list)
        print("Email sent successfully.")
    except Exception as e:
        print(f"Error sending email: {e}")

def test_email(request):
    send_mail(
        'Test Email',
        'This is a test email from your Django project.',
        'your_email@gmail.com',  # Your Gmail address
        ['recipient_email@example.com'],  # The recipient's email address
        fail_silently=False,
    )
    return HttpResponse("Email sent successfully!")


def generate_excel_report(request):
    # Create a workbook and add a worksheet
    workbook = openpyxl.Workbook()
    
    # Create a utility function for styling
    def style_cell(cell, bold=False, fill_color=None, font_color=None):
        if bold:
            cell.font = Font(bold=True)
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        if font_color:
            cell.font = Font(color=font_color)

    # Report for Tours
    tour_sheet = workbook.active
    tour_sheet.title = 'Tour Report'
    
    # Header row
    headers = ['Tour ID', 'Tour Title', 'Total Bookings', 'Total Revenue (₹)', 'Usernames']
    tour_sheet.append(headers)
    
    # Apply header styling
    for col in range(1, len(headers) + 1):
        cell = tour_sheet.cell(row=1, column=col)
        style_cell(cell, bold=True, fill_color='00C0C0C0')  # Light grey fill for header

    # Fetch data for each tour
    tours = Tour.objects.all()
    for tour in tours:
        total_bookings = Booking.objects.filter(tour=tour).count()
        total_revenue = Booking.objects.filter(tour=tour).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        
        # Fetch distinct usernames for the bookings related to this tour
        user_ids = Booking.objects.filter(tour=tour).values_list('user', flat=True).distinct()
        usernames = User.objects.filter(id__in=user_ids).values_list('username', flat=True)
        usernames_list = ', '.join(usernames)  # Join usernames into a single string
        
        tour_sheet.append([tour.id, tour.title, total_bookings, total_revenue, usernames_list])
    
    # Formatting the tour report
    for col in tour_sheet.columns:
        for cell in col:
            cell.alignment = Alignment(horizontal='center')  # Center-align text

    # Create borders for the header
    thin = Side(border_style="thin", color="000000")
    for col in range(1, len(headers) + 1):
        cell = tour_sheet.cell(row=1, column=col)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(2, len(tours) + 2):
        for col in range(1, 6):  # Add borders to data cells
            cell = tour_sheet.cell(row=row, column=col)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Add a summary row at the end of the report
    total_bookings = sum(Booking.objects.filter(tour=tour).count() for tour in tours)
    total_revenue = sum(Booking.objects.filter(tour=tour).aggregate(Sum('total_amount'))['total_amount__sum'] or 0 for tour in tours)
    tour_sheet.append(['', 'Total', total_bookings, total_revenue, ''])

    # Style the summary row
    for col in range(1, 5):
        cell = tour_sheet.cell(row=len(tours) + 2, column=col)
        style_cell(cell, bold=True, fill_color='00FFFF00')  # Yellow fill for summary

    # Create a new worksheet for recent bookings
    recent_sheet = workbook.create_sheet(title='Recent Bookings')
    recent_sheet.append(['Booking ID', 'Tour Title', 'User', 'Amount (₹)', 'Date'])

    # Header styling
    for col in range(1, 5):
        cell = recent_sheet.cell(row=1, column=col)
        style_cell(cell, bold=True, fill_color='00C0C0C0')

    # Fetch the latest 10 bookings and add them to the recent bookings sheet
    recent_bookings = Booking.objects.order_by('-booked_on')[:10]
    for booking in recent_bookings:
        recent_sheet.append([
            booking.id,
            booking.tour.title,
            booking.user.username,
            booking.total_amount,
            booking.booked_on.strftime('%Y-%m-%d'),
        ])

    # Formatting the recent bookings sheet
    for col in recent_sheet.columns:
        for cell in col:
            cell.alignment = Alignment(horizontal='center')  # Center-align text

    # Create a new worksheet for user statistics
    user_sheet = workbook.create_sheet(title='User Report')
    user_sheet.append(['User ID', 'Username', 'Email', 'Total Bookings'])

    # Header styling
    for col in range(1, 5):
        cell = user_sheet.cell(row=1, column=col)
        style_cell(cell, bold=True, fill_color='00C0C0C0')

    # Fetch all users and their total bookings
    users = User.objects.all()
    for user in users:
        total_user_bookings = Booking.objects.filter(user=user).count()
        user_sheet.append([user.id, user.username, user.email, total_user_bookings])

    # Add total users and revenue at the top
    user_sheet.insert_rows(1)
    user_sheet.append(['Total Users', len(users), 'Total Revenue (₹)', 
                       sum(Booking.objects.filter(user=user).aggregate(Sum('total_amount'))['total_amount__sum'] or 0 for user in users)])
    for col in range(1, 5):
        cell = user_sheet.cell(row=1, column=col)
        style_cell(cell, bold=True, fill_color='00FFFF00')  # Yellow fill for summary

    # Formatting the user report
    for col in user_sheet.columns:
        for cell in col:
            cell.alignment = Alignment(horizontal='center')  # Center-align text

    # Save the workbook to a response object
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=project_report.xlsx'
    workbook.save(response)

    return response